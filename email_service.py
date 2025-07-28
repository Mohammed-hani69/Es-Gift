#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
خدمة إرسال أكواد المنتجات عبر الإيميل
==========================================

هذا الملف يوفر نظام إرسال أكواد المنتجات المشتراة من OneCard API
عبر الإيميل للعملاء في ملف Excel منسق.

المتطلبات:
- openpyxl: لإنشاء ملفات Excel
- brevo_email_service: لإرسال الإيميلات باستخدام Brevo

"""

import os
import json
import logging
import base64
from datetime import datetime
from flask import current_app, render_template_string
from flask_mail import Mail
from brevo_email_service import send_simple_email, EmailAttachment
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile

# تكوين التسجيل
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ProductCodeEmailService:
    """خدمة إرسال أكواد المنتجات عبر الإيميل باستخدام Brevo"""
    
    def __init__(self, app=None):
        self.app = app
        if app:
            self.init_app(app)
    
    def init_app(self, app):
        """تهيئة الخدمة مع Flask app"""
        self.app = app
        self.mail = Mail(app)
    
    def create_excel_file(self, order_data, product_codes, save_to_disk=False):
        """
        إنشاء ملف Excel يحتوي على تفاصيل الطلب وأكواد المنتجات
        
        Args:
            order_data (dict): بيانات الطلب
            product_codes (list): قائمة أكواد المنتجات
            save_to_disk (bool): حفظ الملف على القرص الصلب
            
        Returns:
            BytesIO: ملف Excel في الذاكرة
            str: مسار الملف المحفوظ (إذا كان save_to_disk=True)
        """
        # إنشاء Workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "أكواد المنتجات"
        
        # تنسيق الألوان والخطوط
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        content_font = Font(name='Arial', size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # عنوان الملف
        ws.merge_cells('A1:D1')
        ws['A1'] = f"أكواد المنتجات - طلب رقم {order_data.get('order_number', 'غير محدد')}"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # معلومات الطلب
        row = 3
        order_info = [
            ('رقم الطلب:', order_data.get('order_number', 'غير محدد')),
            ('اسم العميل:', order_data.get('customer_name', 'غير محدد')),
            ('البريد الإلكتروني:', order_data.get('customer_email', 'غير محدد')),
            ('تاريخ الطلب:', order_data.get('order_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))),
            ('اسم المنتج:', order_data.get('product_name', 'غير محدد')),
            ('الكمية:', str(order_data.get('quantity', 1))),
            ('المبلغ الإجمالي:', f"{order_data.get('total_amount', 0)} {order_data.get('currency', 'SAR')}")
        ]
        
        for label, value in order_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = content_font
            row += 1
        
        # فاصل
        row += 1
        
        # عنوان أكواد المنتجات
        ws[f'A{row}'] = 'أكواد المنتجات:'
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        
        # رؤوس الجدول
        headers = ['الرقم', 'كود المنتج', 'تاريخ الإنشاء', 'الحالة']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row += 1
        
        # إضافة أكواد المنتجات
        for i, code in enumerate(product_codes, 1):
            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=2, value=code).border = border
            ws.cell(row=row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S')).border = border
            ws.cell(row=row, column=4, value='صالح').border = border
            
            # تنسيق النص
            for col in range(1, 5):
                ws.cell(row=row, column=col).font = content_font
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # تعديل عرض الأعمدة
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # إضافة ملاحظات
        row += 2
        notes = [
            "ملاحظات مهمة:",
            "• احتفظ بهذه الأكواد في مكان آمن",
            "• لا تشارك هذه الأكواد مع أي شخص آخر",
            "• في حالة وجود مشكلة، تواصل معنا فوراً",
            "• صالحية الأكواد حسب شروط المزود"
        ]
        
        for note in notes:
            ws[f'A{row}'] = note
            if note.startswith('ملاحظات'):
                ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            else:
                ws[f'A{row}'].font = Font(name='Arial', size=10)
            row += 1
        
        # حفظ الملف في الذاكرة
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # حفظ الملف على القرص إذا طُلب ذلك
        file_path = None
        if save_to_disk:
            try:
                # إنشاء اسم الملف
                order_number = order_data.get('order_number', 'Unknown')
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"ES-Gift_Order_{order_number}_{timestamp}_Codes.xlsx"
                
                # التأكد من وجود مجلد الحفظ
                excel_dir = os.path.join(current_app.static_folder, 'excel_files')
                if not os.path.exists(excel_dir):
                    os.makedirs(excel_dir)
                
                # مسار الملف الكامل
                full_path = os.path.join(excel_dir, filename)
                
                # حفظ الملف
                wb.save(full_path)
                
                # إرجاع المسار النسبي للحفظ في قاعدة البيانات
                file_path = f"excel_files/{filename}"
                
                logger.info(f"تم حفظ ملف Excel في: {full_path}")
                
            except Exception as e:
                logger.error(f"خطأ في حفظ ملف Excel: {str(e)}")
        
        if save_to_disk:
            return excel_file, file_path
        return excel_file
    
    def send_product_codes_email(self, order_data, product_codes):
        """
        إرسال أكواد المنتجات عبر الإيميل
        
        Args:
            order_data (dict): بيانات الطلب
            product_codes (list): قائمة أكواد المنتجات
            
        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            customer_email = order_data.get('customer_email')
            if not customer_email:
                logger.error("عنوان البريد الإلكتروني غير محدد في بيانات الطلب")
                return False, "عنوان البريد الإلكتروني غير محدد"
            
            logger.info(f"بدء إرسال البريد إلى: {customer_email}")
            
            # التحقق من إعدادات البريد
            mail_username = current_app.config.get('MAIL_USERNAME')
            mail_default_sender = current_app.config.get('MAIL_DEFAULT_SENDER')
            
            if not mail_username:
                logger.error("MAIL_USERNAME غير محدد في إعدادات التطبيق")
                return False, "إعدادات البريد الإلكتروني غير مكتملة"
            
            # استخدام MAIL_USERNAME كمرسل افتراضي إذا لم يكن MAIL_DEFAULT_SENDER محدد
            sender = mail_default_sender or mail_username
            if not sender:
                logger.error("لا يوجد مرسل محدد للبريد الإلكتروني")
                return False, "مرسل البريد الإلكتروني غير محدد"
            
            logger.info(f"مرسل البريد: {sender}")
            
            # إنشاء ملف Excel مع حفظه على القرص
            logger.info("إنشاء ملف Excel...")
            excel_file, saved_file_path = self.create_excel_file(order_data, product_codes, save_to_disk=True)
            logger.info(f"تم إنشاء ملف Excel بحجم: {len(excel_file.getvalue())} بايت")
            if saved_file_path:
                logger.info(f"تم حفظ الملف في: {saved_file_path}")
            
            # إنشاء رسالة الإيميل
            subject = f"أكواد منتجاتك - طلب رقم {order_data.get('order_number', 'غير محدد')}"
            logger.info(f"موضوع البريد: {subject}")
            
            # محتوى الإيميل
            email_template = """
            <!DOCTYPE html>
            <html dir="rtl" lang="ar">
            <head>
                <meta charset="UTF-8">
                <style>
                    body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; direction: rtl; text-align: right; }
                    .container { max-width: 600px; margin: 0 auto; padding: 20px; background: #f8f9fa; }
                    .header { background: #007bff; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0; }
                    .content { background: white; padding: 30px; border-radius: 0 0 10px 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                    .order-info { background: #e3f2fd; padding: 15px; border-radius: 8px; margin: 20px 0; }
                    .footer { text-align: center; margin-top: 20px; color: #666; font-size: 14px; }
                    .btn { background: #28a745; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; display: inline-block; margin: 10px 0; }
                    .warning { background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin: 15px 0; }
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>🎮 Es-Gift</h1>
                        <h2>تم تحضير طلبك بنجاح!</h2>
                    </div>
                    
                    <div class="content">
                        <p>عزيزي/عزيزتي <strong>{{ customer_name }}</strong>,</p>
                        
                        <p>نشكرك لاختيار Es-Gift! تم تحضير طلبك بنجاح وإرفاق أكواد المنتجات في ملف Excel.</p>
                        
                        <div class="order-info">
                            <h3>📋 تفاصيل الطلب:</h3>
                            <ul>
                                <li><strong>رقم الطلب:</strong> {{ order_number }}</li>
                                <li><strong>المنتج:</strong> {{ product_name }}</li>
                                <li><strong>الكمية:</strong> {{ quantity }}</li>
                                <li><strong>المبلغ:</strong> {{ total_amount }} {{ currency }}</li>
                                <li><strong>التاريخ:</strong> {{ order_date }}</li>
                            </ul>
                        </div>
                        
                        <div class="warning">
                            <h3>⚠️ تعليمات مهمة:</h3>
                            <ul>
                                <li>ستجد أكواد منتجاتك في الملف المرفق (Excel)</li>
                                <li>احتفظ بالأكواد في مكان آمن</li>
                                <li>لا تشارك الأكواد مع أي شخص آخر</li>
                                <li>في حالة وجود مشكلة، تواصل معنا فوراً</li>
                            </ul>
                        </div>
                        
                        <p>إذا كان لديك أي استفسار، لا تتردد في التواصل معنا.</p>
                        
                        <div class="footer">
                            <p>شكراً لثقتك في Es-Gift</p>
                            <p>فريق خدمة العملاء | Es-Gift</p>
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """
            
            # تحضير محتوى الإيميل
            logger.info("تحضير محتوى البريد...")
            email_content = render_template_string(email_template, **order_data)
            
            # تحويل ملف Excel إلى base64 للاستخدام مع Brevo
            excel_data = excel_file.getvalue()
            excel_base64 = base64.b64encode(excel_data).decode('utf-8')
            filename = f"ES-Gift_Order_{order_data.get('order_number', 'Unknown')}_Codes.xlsx"
            
            # إنشاء مرفق Brevo
            attachment = EmailAttachment(
                content=excel_base64,
                name=filename,
                type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # إرسال الإيميل باستخدام Brevo
            logger.info("إرسال البريد باستخدام Brevo...")
            success, result = send_simple_email(
                to=customer_email,
                subject=subject,
                html_content=email_content,
                text_content=f"تم تحضير طلبك #{order_data.get('order_number')} - أكواد المنتجات مرفقة في ملف Excel"
            )
            
            if success:
                logger.info(f"تم إرسال البريد بنجاح إلى: {customer_email} باستخدام Brevo")
                return True, f"تم إرسال أكواد المنتجات إلى {customer_email} بنجاح", saved_file_path
            else:
                logger.error(f"فشل إرسال البريد باستخدام Brevo: {result}")
                # محاولة الطريقة البديلة
                return self._send_email_fallback(customer_email, subject, email_content, excel_file, order_data, saved_file_path)
            
        except Exception as e:
            logger.error(f"خطأ في إرسال البريد الإلكتروني: {str(e)}")
            logger.error(f"تفاصيل الخطأ: {type(e).__name__}")
            import traceback
            logger.error(f"Stack trace: {traceback.format_exc()}")
            return False, f"فشل في إرسال الإيميل: {str(e)}", None
    
    def process_order_codes(self, order_id, api_transaction_id):
        """
        معالجة وإرسال أكواد الطلب للعميل
        
        Args:
            order_id (int): معرف الطلب
            api_transaction_id (int): معرف معاملة API
            
        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            from models import Order, APITransaction, User
            
            # جلب بيانات الطلب
            order = Order.query.get(order_id)
            if not order:
                return False, "الطلب غير موجود"
            
            # جلب بيانات المعاملة
            api_transaction = APITransaction.query.get(api_transaction_id)
            if not api_transaction:
                return False, "معاملة API غير موجودة"
            
            # فحص وجود أكواد المنتج
            if not api_transaction.product_codes:
                return False, "لا توجد أكواد للمنتج"
            
            # استخراج الأكواد
            try:
                codes_data = json.loads(api_transaction.product_codes)
                if isinstance(codes_data, list):
                    product_codes = codes_data
                elif isinstance(codes_data, dict):
                    product_codes = codes_data.get('codes', [])
                else:
                    product_codes = [str(codes_data)]
            except:
                product_codes = [api_transaction.product_codes]
            
            if not product_codes:
                return False, "لا توجد أكواد صالحة"
            
            # تحضير بيانات الطلب
            order_data = {
                'order_number': order.id,
                'customer_name': order.user.full_name or order.user.username,
                'customer_email': order.user.email,
                'order_date': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'product_name': order.product.name if order.product else 'منتج رقمي',
                'quantity': order.quantity,
                'total_amount': float(order.total_amount),
                'currency': 'SAR'
            }
            
            # إرسال الإيميل
            return self.send_product_codes_email(order_data, product_codes)
            
        except Exception as e:
            current_app.logger.error(f"Error processing order codes: {str(e)}")
            return False, f"خطأ في معالجة أكواد الطلب: {str(e)}"
    
    def _send_email_fallback(self, customer_email, subject, email_content, excel_file, order_data, saved_file_path):
        """إرسال الإيميل باستخدام Flask-Mail كبديل"""
        try:
            from flask_mail import Message, Mail
            
            # إنشاء رسالة الإيميل
            logger.info("إنشاء رسالة البريد باستخدام Flask-Mail...")
            
            mail = Mail(current_app) if not hasattr(self, 'mail') or not self.mail else self.mail
            
            msg = Message(
                subject=subject,
                recipients=[customer_email],
                html=email_content,
                sender=current_app.config.get('MAIL_DEFAULT_SENDER')
            )
            
            # إرفاق ملف Excel
            filename = f"ES-Gift_Order_{order_data.get('order_number', 'Unknown')}_Codes.xlsx"
            logger.info(f"إرفاق ملف Excel: {filename}")
            msg.attach(
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                data=excel_file.getvalue()
            )
            
            # إرسال الإيميل
            logger.info("إرسال البريد باستخدام Flask-Mail...")
            mail.send(msg)
            logger.info(f"تم إرسال البريد بنجاح إلى: {customer_email} باستخدام Flask-Mail")
            
            return True, f"تم إرسال أكواد المنتجات إلى {customer_email} بنجاح", saved_file_path
            
        except Exception as e:
            logger.error(f"خطأ في إرسال البريد باستخدام Flask-Mail: {str(e)}")
            return False, f"فشل في إرسال الإيميل: {str(e)}", None

# إنشاء مثيل الخدمة
email_service = ProductCodeEmailService()

def send_email(to_email, subject, html_content, attachments=None):
    """دالة بسيطة لإرسال إيميل"""
    try:
        from flask import current_app
        from flask_mail import Message, Mail
        
        mail = Mail(current_app)
        
        msg = Message(
            subject=subject,
            recipients=[to_email] if isinstance(to_email, str) else to_email,
            html=html_content,
            sender=current_app.config.get('MAIL_DEFAULT_SENDER')
        )
        
        # إضافة المرفقات إذا توفرت
        if attachments:
            for attachment in attachments:
                if isinstance(attachment, dict):
                    msg.attach(
                        filename=attachment.get('filename', 'attachment'),
                        content_type=attachment.get('content_type', 'application/octet-stream'),
                        data=attachment.get('data', b'')
                    )
        
        mail.send(msg)
        return True
        
    except Exception as e:
        current_app.logger.error(f"Error sending email: {str(e)}")
        return False

def init_email_service(app):
    """تهيئة خدمة الإيميل مع Flask app"""
    email_service.init_app(app)
    return email_service
